#!/usr/bin/env python3
"""
build_gl_pivot_html.py

Reads all TIM GL Data CSV files (semicolon-delimited, UTF-8), aggregates
the data by key business dimensions, enriches with OWC accounts.xlsx and
SEGMENTS.xlsx lookups, then writes a self-contained interactive HTML pivot
table (gl_pivot.html) you can open directly in any browser.

Enrichment columns added automatically
---------------------------------------
From OWC accounts.xlsx (joined on Conto Co.Ge. = Numero conto):
  OWC Testo       — Balance sheet / P&L position text
  OWC Categoria   — 8 top-level OWC categories (Inventories, Trade Receivable …)
  OWC Codice DB   — Detailed DB code (55 codes)

From SEGMENTS.xlsx CDC sheet:
  --segment-summary  Resolves Segmento → SEGMENTI + FUNZIONE inside each chunk,
                     then groups by those (5 × 31 values). Compact output for
                     all 6 months combined — recommended default with segments.
  --with-segment     Groups by raw Segmento code — ~14× more rows per file.
                     Use only for single-month deep drill-downs.

Usage
-----
    python3 build_gl_pivot_html.py                        # all months, OWC only
    python3 build_gl_pivot_html.py --segment-summary      # all months + SEGMENTI/FUNZIONE (recommended)
    python3 build_gl_pivot_html.py --output my.html       # custom output name
    python3 build_gl_pivot_html.py --chunk 200000         # larger chunks if RAM allows
    python3 build_gl_pivot_html.py --with-segment         # raw Segmento grouping (single month recommended)
    python3 build_gl_pivot_html.py --months Jul Dec       # specific months only

CSV column layout (semicolon separator)
----------------------------------------
 0  Num.Doc.        Document number
 1  Pos.doc.        Line position
 2  Conto Co.Ge.    GL account code
 3  Conto sez.      Section account
 4  Società         Company code
 5  Importo         Dare / Debit amount
 6  Imp. Avere      Avere / Credit amount
 7  Div. Soc.       Currency
 8  Eser.           Fiscal year
 9  Per.            Fiscal period (1–12)
10  Data reg.       Registration date
11  Data acq.       Acquisition date
12  User            SAP user ID
13  Tipo doc.       Document type
14  testo testata   Header text
15  Testo Pos.      Line item text
16  Data doc.       Document date
17  Segmento        Segment / profit-centre code (no header in file)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ── constants ─────────────────────────────────────────────────────────────────

MONTH_NAMES: dict[int, str] = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
    13: "Adj",   # SAP period 13 — year-end adjustment postings
}
MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec","Adj"]

# Full 18-column name list; the file header only has 17 names, col 17 is unnamed
CSV_ALL_COLS = [
    "Num.Doc.", "Pos.doc.", "Conto Co.Ge.", "Conto sez.", "Società",
    "Importo", "Imp. Avere", "Div. Soc.", "Eser.", "Per.",
    "Data reg.", "Data acq.", "User", "Tipo doc.", "testo testata",
    "Testo Pos.", "Data doc.", "Segmento",
]

NEEDED_COLS_BASE = [
    "Conto Co.Ge.", "Società", "Importo", "Imp. Avere",
    "Div. Soc.", "Eser.", "Per.", "Tipo doc.",
]
NEEDED_COLS_SEG = NEEDED_COLS_BASE + ["Segmento"]

BASE_GROUP_KEYS    = ["Società", "Anno", "Mese", "Conto Co.Ge.", "Tipo doc.", "Div. Soc."]
FULL_GROUP_KEYS    = BASE_GROUP_KEYS + ["Segmento"]
SUMMARY_GROUP_KEYS = BASE_GROUP_KEYS + ["SEGMENTI", "FUNZIONE"]  # compact segment mode

_NONE_STR = "(vuoto)"


# ── reference-file loaders ────────────────────────────────────────────────────

def _norm(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s.upper()


def load_owc_lookup(path: Path) -> pd.DataFrame:
    """
    Returns a DataFrame indexed by normalised GL account code with columns:
      OWC Testo, OWC Categoria, OWC Codice DB
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Tabella OWC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Header is at index 3 (rows 0-2 are blank)
    records = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        code = _norm(row[0])
        if not code:
            continue
        records.append({
            "Conto Co.Ge.":  code,
            "OWC Testo":     str(row[1]).strip() if row[1] else "",
            "OWC Categoria": str(row[2]).strip() if row[2] else "(Non-OWC)",
            "OWC Codice DB": str(row[3]).strip() if row[3] else "",
        })

    df = pd.DataFrame(records).drop_duplicates("Conto Co.Ge.").set_index("Conto Co.Ge.")
    print(f"  OWC accounts loaded: {len(df):,} GL codes, "
          f"{df['OWC Categoria'].nunique()} categories")
    return df


def load_segments_lookup(path: Path) -> tuple[pd.DataFrame, dict]:
    """
    Returns (DataFrame, dict) where:
      DataFrame — indexed by segment code, columns: Seg Descrizione, SEGMENTI, FUNZIONE
      dict      — {code: (SEGMENTI, FUNZIONE)} for fast per-row chunk lookups
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["CDC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    fast_lookup: dict[str, tuple[str, str]] = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip().upper()
        segmenti = str(row[6]).strip() if row[6] else "(Unmapped)"
        funzione = str(row[7]).strip() if row[7] else "(Unmapped)"
        records.append({
            "Segmento":        code,
            "Seg Descrizione": str(row[1]).strip() if row[1] else "",
            "SEGMENTI":        segmenti,
            "FUNZIONE":        funzione,
        })
        fast_lookup[code] = (segmenti, funzione)

    df = pd.DataFrame(records).drop_duplicates("Segmento").set_index("Segmento")
    print(f"  SEGMENTS (CDC) loaded: {len(df):,} codes, "
          f"{df['SEGMENTI'].nunique()} top-level segments, "
          f"{df['FUNZIONE'].nunique()} functions")
    return df, fast_lookup


# ── helpers ───────────────────────────────────────────────────────────────────

def _parse_amounts(series: pd.Series) -> pd.Series:
    """Coerce '0.00 ' / '19.98 ' style strings to float."""
    return pd.to_numeric(series.astype(str).str.strip(), errors="coerce").fillna(0.0)


def _month_label(per_val: str) -> str:
    try:
        return MONTH_NAMES.get(int(per_val), per_val)
    except (ValueError, TypeError):
        return str(per_val)


# ── per-file CSV processing ───────────────────────────────────────────────────

def process_file(
    csv_path: Path,
    group_keys: list[str],
    needed_cols: list[str],
    chunk_size: int,
    seg_lookup: dict | None = None,   # {code: (SEGMENTI, FUNZIONE)} for summary mode
) -> pd.DataFrame:
    """Read *csv_path* in chunks, aggregate within each chunk, return combined df."""
    parts: list[pd.DataFrame] = []
    t0 = time.time()
    total_rows = 0
    use_seg_summary = seg_lookup is not None

    for chunk_idx, chunk in enumerate(
        pd.read_csv(
            csv_path,
            sep=";",
            header=0,
            names=CSV_ALL_COLS,
            usecols=needed_cols,
            dtype=str,
            encoding="utf-8",
            encoding_errors="replace",
            chunksize=chunk_size,
            on_bad_lines="skip",
            low_memory=False,
        ),
        start=1,
    ):
        total_rows += len(chunk)
        if chunk_idx % 20 == 0:
            elapsed = time.time() - t0
            print(
                f"    chunk {chunk_idx:>4}  |  {total_rows:>12,} rows  |  {elapsed:.0f}s",
                flush=True,
            )

        chunk["Importo"]    = _parse_amounts(chunk["Importo"])
        chunk["Imp. Avere"] = _parse_amounts(chunk["Imp. Avere"])

        str_cols = ["Conto Co.Ge.", "Società", "Div. Soc.", "Eser.", "Per.", "Tipo doc."]
        if "Segmento" in chunk.columns:
            str_cols.append("Segmento")
        for col in str_cols:
            chunk[col] = chunk[col].fillna("").astype(str).str.strip()

        chunk["Anno"] = chunk["Eser."]
        chunk["Mese"] = chunk["Per."].map(_month_label)

        # Segment-summary mode: resolve raw code → SEGMENTI / FUNZIONE inline
        if use_seg_summary and "Segmento" in chunk.columns:
            seg_col = chunk["Segmento"].str.upper()
            chunk["SEGMENTI"] = seg_col.map(lambda c: seg_lookup.get(c, ("(Unmapped)", "(Unmapped)"))[0])
            chunk["FUNZIONE"]  = seg_col.map(lambda c: seg_lookup.get(c, ("(Unmapped)", "(Unmapped)"))[1])

        agg = (
            chunk.groupby(group_keys, dropna=False, sort=False)
            .agg(
                Importo=("Importo", "sum"),
                Imp_Avere=("Imp. Avere", "sum"),
                Transazioni=("Importo", "count"),
            )
            .reset_index()
        )
        parts.append(agg)

    if not parts:
        return pd.DataFrame()

    combined = pd.concat(parts, ignore_index=True)
    result = (
        combined.groupby(group_keys, dropna=False, sort=False)
        .agg(
            Importo=("Importo", "sum"),
            Imp_Avere=("Imp_Avere", "sum"),
            Transazioni=("Transazioni", "sum"),
        )
        .reset_index()
    )
    elapsed = time.time() - t0
    print(
        f"  Done: {total_rows:,} rows → {len(result):,} aggregated  ({elapsed:.0f}s)",
        flush=True,
    )
    return result


# ── enrichment join ───────────────────────────────────────────────────────────

def enrich(
    df: pd.DataFrame,
    owc: pd.DataFrame,
    segments: pd.DataFrame | None,
) -> pd.DataFrame:
    """Left-join OWC and (optionally) SEGMENTS onto the aggregated GL data."""
    # Normalise key for join
    df["_key"] = df["Conto Co.Ge."].str.upper()
    df = df.join(owc, on="_key", how="left")
    df.drop(columns=["_key"], inplace=True)

    # Fill unmapped accounts
    df["OWC Testo"]     = df["OWC Testo"].fillna("(Non-OWC)")
    df["OWC Categoria"] = df["OWC Categoria"].fillna("(Non-OWC)")
    df["OWC Codice DB"] = df["OWC Codice DB"].fillna("(Non-OWC)")

    # Full raw-segment join (--with-segment mode)
    if segments is not None and "Segmento" in df.columns and "SEGMENTI" not in df.columns:
        df["_seg"] = df["Segmento"].str.upper()
        df = df.join(segments, on="_seg", how="left")
        df.drop(columns=["_seg"], inplace=True)
        df["Seg Descrizione"] = df["Seg Descrizione"].fillna("(Unmapped)")
        df["SEGMENTI"]        = df["SEGMENTI"].fillna("(Unmapped)")
        df["FUNZIONE"]        = df["FUNZIONE"].fillna("(Unmapped)")

    return df


# ── HTML builder ──────────────────────────────────────────────────────────────

def build_html(
    records: list[dict],
    num_files: int,
    has_owc: bool,
    has_segments: bool,
    seg_summary: bool = False,   # True when SEGMENTI/FUNZIONE are group keys (no raw Segmento)
) -> str:
    data_json   = json.dumps(records, ensure_ascii=False, default=str)
    num_records = len(records)

    owc_options = ""
    owc_presets = ""
    if has_owc:
        owc_options = """
    <optgroup label="OWC Accounts">
      <option value="owc_categoria_mese">OWC Categoria by Month (Importo)</option>
      <option value="owc_categoria_societa">OWC Categoria by Company (Importo)</option>
      <option value="owc_db_mese">OWC Codice DB by Month (Importo)</option>
      <option value="owc_testo_mese">OWC Balance-Sheet Position by Month</option>
      <option value="owc_net_categoria">Net (Dare–Avere) by OWC Categoria &amp; Company (heatmap)</option>
    </optgroup>"""
        owc_presets = """
  owc_categoria_mese: {
    rows: ["OWC Categoria"],
    cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },
  owc_categoria_societa: {
    rows: ["OWC Categoria"],
    cols: ["Società"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },
  owc_db_mese: {
    rows: ["OWC Codice DB"],
    cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },
  owc_testo_mese: {
    rows: ["OWC Testo"],
    cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },
  owc_net_categoria: {
    rows: ["OWC Categoria"],
    cols: ["Società"],
    aggName: "Sum", vals: ["Netto"], renderer: "Table Heatmap"
  },"""

    seg_options = ""
    seg_presets = ""
    if has_segments:
        raw_option  = "" if seg_summary else '<option value="seg_raw_mese">Raw Segmento code by Month</option>'
        desc_option = "" if seg_summary else '<option value="seg_desc_mese">Segment Description by Month</option>'
        seg_options = f"""
    <optgroup label="Segments (CDC)">
      <option value="segmenti_mese">SEGMENTI by Month (Importo)</option>
      <option value="segmenti_societa">SEGMENTI by Company (Importo)</option>
      <option value="funzione_mese">FUNZIONE by Month (Importo)</option>
      <option value="segmenti_funzione_mese">SEGMENTI + FUNZIONE by Month</option>
      <option value="netto_segmenti">Net by SEGMENTI &amp; Company (heatmap)</option>
      {desc_option}
      {raw_option}
    </optgroup>"""
        raw_preset  = "" if seg_summary else """
  seg_raw_mese: {
    rows: ["Segmento"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },"""
        desc_preset = "" if seg_summary else """
  seg_desc_mese: {
    rows: ["Seg Descrizione"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  },"""
        seg_presets = f"""
  segmenti_mese: {{
    rows: ["SEGMENTI"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  segmenti_societa: {{
    rows: ["SEGMENTI"], cols: ["Società"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  funzione_mese: {{
    rows: ["FUNZIONE"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  segmenti_funzione_mese: {{
    rows: ["SEGMENTI", "FUNZIONE"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  netto_segmenti: {{
    rows: ["SEGMENTI"], cols: ["Società"],
    aggName: "Sum", vals: ["Netto"], renderer: "Table Heatmap"
  }},{raw_preset}{desc_preset}"""

    enrichment_badge = ""
    if has_owc:
        enrichment_badge += " &nbsp;·&nbsp; OWC enriched"
    if has_segments:
        enrichment_badge += " &nbsp;·&nbsp; SEGMENTS enriched"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TIM GL Data — Pivot Table</title>

  <link rel="stylesheet"
    href="https://cdnjs.cloudflare.com/ajax/libs/jqueryui/1.13.3/themes/cupertino/jquery-ui.min.css"
    crossorigin="anonymous">
  <link rel="stylesheet"
    href="https://cdnjs.cloudflare.com/ajax/libs/pivottable/2.23.0/pivot.min.css"
    crossorigin="anonymous">

  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

    body {{
      font-family: system-ui, -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
      background: #f0f2f7;
      color: #1b2230;
      min-height: 100vh;
    }}

    header {{
      background: #162d50;
      color: #fff;
      padding: 12px 24px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
    }}
    header h1 {{ font-size: 1.1rem; font-weight: 600; letter-spacing: .01em; }}
    .header-meta {{ font-size: 0.78rem; opacity: .65; white-space: nowrap; }}

    .toolbar {{
      background: #fff;
      border-bottom: 1px solid #dde4ef;
      padding: 9px 24px;
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
    }}
    .toolbar-label {{
      font-size: 0.8rem; font-weight: 600; color: #6b7a90; white-space: nowrap;
    }}
    .toolbar select {{
      font-size: 0.82rem; padding: 5px 9px; border: 1px solid #c8d1e0;
      border-radius: 5px; background: #fff; color: #1b2230;
      cursor: pointer; min-width: 260px;
    }}
    .toolbar button {{
      font-size: 0.82rem; font-weight: 600; padding: 5px 14px;
      border: none; border-radius: 5px; background: #162d50;
      color: #fff; cursor: pointer; white-space: nowrap;
    }}
    .toolbar button:hover {{ background: #1f3f6e; }}
    .toolbar-hint {{
      margin-left: auto; font-size: 0.76rem; color: #8a95a8; white-space: nowrap;
    }}

    .statsbar {{
      background: #fff;
      border-bottom: 1px solid #dde4ef;
      padding: 7px 24px;
      display: flex;
      gap: 22px;
      font-size: 0.79rem;
      color: #5f6b7d;
      flex-wrap: wrap;
    }}
    .statsbar strong {{ color: #1b2230; }}
    .statsbar .sep {{
      border-left: 1px solid #dde4ef;
      padding-left: 22px;
    }}

    #pivot-wrap {{ padding: 20px 24px; overflow: auto; }}

    .pvtUi {{
      background: #fff; border: 1px solid #dde4ef;
      border-radius: 8px; padding: 16px; font-size: 0.83rem;
    }}
    .pvtTable {{ font-size: 0.82rem; }}
    .pvtTable thead tr th {{
      background: #eef2fb; font-weight: 600; white-space: nowrap;
    }}
    .pvtTotal, .pvtGrandTotal {{
      background: #f5f7fc !important; font-weight: 600;
    }}
    .pvtAxisContainer {{
      border: 1px solid #dde4ef !important;
      border-radius: 6px !important;
      background: #f8fafd !important;
    }}
    .pvtFilterBox {{ max-height: 360px; }}

    footer {{
      text-align: center; padding: 12px;
      font-size: 0.73rem; color: #9aa5b4;
    }}
  </style>
</head>
<body>

<header>
  <h1>TIM GL Data — Interactive Pivot Table</h1>
  <span class="header-meta">
    {num_files} monthly file{'s' if num_files != 1 else ''}
    &nbsp;·&nbsp; {num_records:,} aggregated rows{enrichment_badge}
  </span>
</header>

<div class="toolbar">
  <span class="toolbar-label">Quick preset:</span>
  <select id="preset-select">
    <optgroup label="GL Core">
      <option value="importo_societa">Importo (Dare) by Company &amp; Month</option>
      <option value="avere_societa">Imp. Avere (Credit) by Company &amp; Month</option>
      <option value="netto_societa">Net (Dare–Avere) by Company &amp; Month</option>
      <option value="importo_conto">Importo by GL Account &amp; Month</option>
      <option value="netto_conto">Net by GL Account &amp; Company (heatmap)</option>
      <option value="tx_doctype">Transactions by Doc Type &amp; Company (heatmap)</option>
      <option value="importo_anno_mese">Importo by Year/Month &amp; Company</option>
    </optgroup>
    {owc_options}
    {seg_options}
  </select>
  <button id="btn-apply">Apply</button>
  <span class="toolbar-hint">Drag any field · Click cells to filter · Change aggregator in UI</span>
</div>

<div class="statsbar" id="statsbar">
  <span>Loading data&hellip;</span>
</div>

<div id="pivot-wrap">
  <div id="pivot-output"></div>
</div>

<footer>
  Importo = Dare (debit); Imp. Avere = credit; Netto = Importo &minus; Imp.&nbsp;Avere.
  Amounts in document currency (Div. Soc.). &nbsp;|&nbsp; Built with PivotTable.js.
</footer>

<script src="https://cdnjs.cloudflare.com/ajax/libs/jquery/3.7.1/jquery.min.js"
  crossorigin="anonymous"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jqueryui/1.13.3/jquery-ui.min.js"
  crossorigin="anonymous"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/pivottable/2.23.0/pivot.min.js"
  crossorigin="anonymous"></script>

<script>
var GL_DATA = {data_json};

var PRESETS = {{
  importo_societa: {{
    rows: ["Società"], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  avere_societa: {{
    rows: ["Società"], cols: ["Mese"],
    aggName: "Sum", vals: ["Imp. Avere"], renderer: "Table"
  }},
  netto_societa: {{
    rows: ["Società"], cols: ["Mese"],
    aggName: "Sum", vals: ["Netto"], renderer: "Table"
  }},
  importo_conto: {{
    rows: ["Conto Co.Ge.", "Tipo doc."], cols: ["Mese"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  netto_conto: {{
    rows: ["Conto Co.Ge."], cols: ["Società"],
    aggName: "Sum", vals: ["Netto"], renderer: "Table Heatmap"
  }},
  tx_doctype: {{
    rows: ["Tipo doc."], cols: ["Società"],
    aggName: "Sum", vals: ["Transazioni"], renderer: "Table Heatmap"
  }},
  importo_anno_mese: {{
    rows: ["Anno", "Mese"], cols: ["Società"],
    aggName: "Sum", vals: ["Importo"], renderer: "Table"
  }},
  {owc_presets}
  {seg_presets}
}};

// ── stats bar ────────────────────────────────────────────────────────────────
(function() {{
  var companies = {{}}, months = {{}}, accounts = {{}}, currencies = {{}};
  var owcCats = {{}}, segBuckets = {{}};
  var totalImporto = 0, totalAvere = 0;
  GL_DATA.forEach(function(r) {{
    companies[r["Società"]]     = 1;
    months[r["Mese"]]           = 1;
    accounts[r["Conto Co.Ge."]] = 1;
    currencies[r["Div. Soc."]]  = 1;
    if (r["OWC Categoria"]) owcCats[r["OWC Categoria"]] = 1;
    if (r["SEGMENTI"])      segBuckets[r["SEGMENTI"]]   = 1;
    totalImporto += (r["Importo"]    || 0);
    totalAvere   += (r["Imp. Avere"] || 0);
  }});
  var fmt = function(n) {{
    return n.toLocaleString("it-IT", {{minimumFractionDigits: 0, maximumFractionDigits: 0}});
  }};
  var html =
    "<span>Companies: <strong>" + Object.keys(companies).length + "</strong></span>" +
    "<span>GL Accounts: <strong>" + Object.keys(accounts).length + "</strong></span>" +
    "<span>Months: <strong>" + Object.keys(months).length + "</strong></span>" +
    "<span>Currencies: <strong>" + Object.keys(currencies).join(", ") + "</strong></span>" +
    "<span>Total Importo: <strong>" + fmt(totalImporto) + "</strong></span>" +
    "<span>Total Avere: <strong>" + fmt(totalAvere) + "</strong></span>";
  if (Object.keys(owcCats).length)
    html += "<span class='sep'>OWC Categories: <strong>" + Object.keys(owcCats).length + "</strong></span>";
  if (Object.keys(segBuckets).length)
    html += "<span>Segments: <strong>" + Object.keys(segBuckets).join(", ") + "</strong></span>";
  $("#statsbar").html(html);
}})();

// ── month sort ───────────────────────────────────────────────────────────────
var MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec","Adj"];
var monthSorter = $.pivotUtilities.sortAs(MONTH_ORDER);

// ── render ───────────────────────────────────────────────────────────────────
function renderPreset(key) {{
  var p = PRESETS[key] || PRESETS["importo_societa"];
  $("#pivot-output").pivotUI(GL_DATA, {{
    rows: p.rows || [],
    cols: p.cols || [],
    aggregatorName: p.aggName || "Sum",
    vals: p.vals || [],
    rendererName: p.renderer || "Table",
    sorters: {{ "Mese": monthSorter }},
    menuLimit: 5000,
    unusedAttrsVertical: true,
  }}, /*overwrite=*/true);
}}

$(function() {{
  renderPreset("importo_societa");
  $("#btn-apply").on("click", function() {{
    renderPreset($("#preset-select").val());
  }});
}});
</script>
</body>
</html>
"""


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Build an interactive pivot-table HTML from TIM GL Data CSVs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument(
        "--root", type=Path, default=Path(__file__).resolve().parent,
        help="Directory containing 'TIM GL Data/', OWC accounts.xlsx, SEGMENTS.xlsx",
    )
    ap.add_argument(
        "--output", type=Path, default=Path("gl_pivot.html"),
        help="Output HTML file (default: gl_pivot.html next to script)",
    )
    ap.add_argument(
        "--chunk", type=int, default=100_000, metavar="N",
        help="Rows per CSV chunk (default 100 000); increase if RAM allows",
    )
    ap.add_argument(
        "--segment-summary", action="store_true", dest="seg_summary",
        help=(
            "Recommended: resolve Segmento → SEGMENTI + FUNZIONE inside each chunk "
            "and group by those. Adds segment dimensions with compact output size — "
            "works well for all 6 months combined."
        ),
    )
    ap.add_argument(
        "--with-segment", action="store_true", dest="with_segment",
        help=(
            "Group by raw Segmento code AND join SEGMENTS.xlsx for full descriptions. "
            "Produces ~14× more rows per file — recommended for single-month runs only."
        ),
    )
    ap.add_argument(
        "--no-owc", action="store_true",
        help="Skip OWC accounts.xlsx enrichment",
    )
    ap.add_argument(
        "--months", nargs="+", metavar="MONTH",
        help="Process only these months (e.g. Jul Aug Dec). Case-insensitive.",
    )
    # legacy alias
    ap.add_argument("--no-segment", action="store_true", help=argparse.SUPPRESS)
    args = ap.parse_args()

    root    = args.root.resolve()
    gl_dir  = root / "TIM GL Data"
    csv_files = sorted(gl_dir.glob("*.csv"))

    if not csv_files:
        print(f"ERROR: No CSV files found in {gl_dir}", file=sys.stderr)
        return 1

    if args.months:
        wanted    = {m.lower() for m in args.months}
        csv_files = [p for p in csv_files
                     if any(m.lower() in p.stem.lower() for m in wanted)]
        if not csv_files:
            print("ERROR: No CSV files matched --months filter.", file=sys.stderr)
            return 1

    use_segment  = getattr(args, "with_segment", False)
    use_seg_sum  = getattr(args, "seg_summary", False) and not use_segment
    use_owc      = not getattr(args, "no_owc", False)

    if use_segment:
        group_keys  = FULL_GROUP_KEYS
        needed_cols = NEEDED_COLS_SEG
        seg_mode    = "raw Segmento (--with-segment)"
    elif use_seg_sum:
        group_keys  = SUMMARY_GROUP_KEYS
        needed_cols = NEEDED_COLS_SEG
        seg_mode    = "SEGMENTI+FUNZIONE summary (--segment-summary)"
    else:
        group_keys  = BASE_GROUP_KEYS
        needed_cols = NEEDED_COLS_BASE
        seg_mode    = "none  (add --segment-summary to include segment dimensions)"

    print("TIM GL Pivot Builder")
    print(f"  Source dir : {gl_dir}")
    print(f"  CSV files  : {len(csv_files)}")
    print(f"  Chunk size : {args.chunk:,}")
    print(f"  OWC enrich : {'yes' if use_owc else 'no (--no-owc)'}")
    print(f"  Segments   : {seg_mode}")
    print()

    # ── load reference data ──────────────────────────────────────────────────
    owc_df      = None
    seg_df      = None
    seg_lookup  = None   # fast dict for inline summary mode

    if use_owc:
        owc_path = root / "OWC accounts.xlsx"
        if owc_path.exists():
            print("Loading OWC accounts.xlsx…")
            owc_df = load_owc_lookup(owc_path)
        else:
            print(f"WARNING: OWC accounts.xlsx not found at {owc_path} — skipping OWC enrichment")

    if use_segment or use_seg_sum:
        seg_path = root / "SEGMENTS.xlsx"
        if seg_path.exists():
            print("Loading SEGMENTS.xlsx…")
            seg_df, seg_lookup = load_segments_lookup(seg_path)
        else:
            print(f"WARNING: SEGMENTS.xlsx not found at {seg_path} — skipping segment enrichment")
            seg_lookup = None
    print()

    # ── process CSV files ────────────────────────────────────────────────────
    all_parts: list[pd.DataFrame] = []
    for i, csv_path in enumerate(csv_files, start=1):
        print(f"[{i}/{len(csv_files)}] {csv_path.name}")
        # Pass seg_lookup only in summary mode (inline resolution per chunk)
        df = process_file(
            csv_path, group_keys, needed_cols, args.chunk,
            seg_lookup=seg_lookup if use_seg_sum else None,
        )
        if not df.empty:
            all_parts.append(df)
        print()

    if not all_parts:
        print("ERROR: No data was aggregated.", file=sys.stderr)
        return 1

    # ── combine ──────────────────────────────────────────────────────────────
    print("Combining all files…", flush=True)
    combined = pd.concat(all_parts, ignore_index=True)
    final = (
        combined.groupby(group_keys, dropna=False, sort=False)
        .agg(
            Importo=("Importo", "sum"),
            Imp_Avere=("Imp_Avere", "sum"),
            Transazioni=("Transazioni", "sum"),
        )
        .reset_index()
    )
    final["Netto"]      = (final["Importo"] - final["Imp_Avere"]).round(2)
    final["Importo"]    = final["Importo"].round(2)
    final["Imp_Avere"]  = final["Imp_Avere"].round(2)
    final.rename(columns={"Imp_Avere": "Imp. Avere"}, inplace=True)

    # Calendar sort
    month_order_map = {m: i for i, m in enumerate(MONTH_ORDER)}
    final["_ms"] = final["Mese"].map(month_order_map).fillna(99)
    final.sort_values(["Anno", "_ms", "Società", "Conto Co.Ge."], inplace=True)
    final.drop(columns=["_ms"], inplace=True)

    # ── enrich ───────────────────────────────────────────────────────────────
    if owc_df is not None or seg_df is not None:
        print("Enriching with reference data…", flush=True)
        final = enrich(final, owc_df if owc_df is not None else pd.DataFrame(), seg_df)
        if owc_df is not None:
            mapped   = (final["OWC Categoria"] != "(Non-OWC)").sum()
            total    = len(final)
            print(f"  OWC coverage: {mapped:,}/{total:,} rows ({100*mapped/total:.1f}%)")
        if seg_df is not None and "SEGMENTI" in final.columns:
            mapped = (final["SEGMENTI"] != "(Unmapped)").sum()
            total  = len(final)
            print(f"  SEGMENTS coverage: {mapped:,}/{total:,} rows ({100*mapped/total:.1f}%)")

    records  = final.to_dict(orient="records")
    has_owc  = owc_df is not None
    has_segs = "SEGMENTI" in final.columns
    print(f"\nFinal aggregated rows: {len(records):,}")

    size_kb = len(json.dumps(records, default=str).encode()) / 1024
    print(f"Estimated JSON size  : {size_kb:,.0f} KB")
    if size_kb > 50_000:
        print(
            "WARNING: JSON payload > 50 MB — browser may be slow.\n"
            "         Consider using --segment-summary instead of --with-segment."
        )

    html_content = build_html(
        records, len(csv_files), has_owc, has_segs,
        seg_summary=use_seg_sum,
    )

    out = args.output if args.output.is_absolute() else root / args.output
    out.write_text(html_content, encoding="utf-8")
    out_kb = out.stat().st_size / 1024
    print(f"\nWrote: {out}  ({out_kb:,.0f} KB)")
    print("Open gl_pivot.html in your browser — no server needed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
